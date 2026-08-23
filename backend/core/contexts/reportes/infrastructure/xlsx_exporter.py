from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class XlsxReportExporter:
    """Adaptador XLSX: recibe datos tabulares, sin conocer reglas del reporte."""

    def export(self, *, title, headers, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title[:31]
        sheet.freeze_panes = "A2"
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="064AA0")
        for row in rows:
            sheet.append(list(row))
        for column_index, header in enumerate(headers, start=1):
            values = [str(sheet.cell(row=index, column=column_index).value or "") for index in range(1, sheet.max_row + 1)]
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(len(header) + 2, *(len(value) + 2 for value in values)), 42)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
