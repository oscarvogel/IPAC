from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase
from django.core.management.base import CommandError
from django.utils import timezone
from decimal import Decimal
from datetime import timedelta
from rest_framework import status
from rest_framework.test import APIClient, APITestCase
from unittest.mock import patch
from io import BytesIO
from openpyxl import load_workbook

from .contexts.importacion.application.import_ipac_workbook import IPACWorkbookImporter, parse_date_value, split_full_name
from .contexts.cobranzas.application.registrar_pago import RegistrarPago
from .models import AplicacionPago, Alumno, CajaDiaria, CarreraCurso, ConceptoCobrable, Cuota, EventoAuditoria, Matricula, MovimientoCaja, Pago, PerfilUsuario, ReglaRecargo, SaldoArrastrableCaja, Sucursal, TipoDescuento


class SucursalSeedTests(TestCase):
    def test_seed_creates_posadas_and_eldorado(self):
        call_command("seed_initial_data", verbosity=0)

        self.assertTrue(Sucursal.objects.filter(codigo="POS", nombre="Posadas").exists())
        self.assertTrue(Sucursal.objects.filter(codigo="ELD", nombre="Eldorado").exists())

    def test_seed_creates_initial_admin_with_profile(self):
        call_command("seed_initial_data", verbosity=0)

        admin = User.objects.get(username="admin")

        self.assertTrue(admin.is_staff)
        self.assertEqual(admin.perfil.rol, PerfilUsuario.Rol.ADMINISTRACION)
        self.assertTrue(admin.perfil.puede_ver_todas_las_sucursales)

    def test_seed_creates_demo_catalogs_and_students(self):
        call_command("seed_initial_data", verbosity=0)

        self.assertGreaterEqual(CarreraCurso.objects.count(), 2)
        self.assertGreaterEqual(ConceptoCobrable.objects.count(), 2)
        self.assertGreaterEqual(Alumno.objects.count(), 3)

    @patch.dict("os.environ", {"IPAC_ENVIRONMENT": "production", "IPAC_SEED_ADMIN_USERNAME": "", "IPAC_SEED_ADMIN_PASSWORD": ""})
    def test_production_seed_requires_explicit_non_demo_credentials(self):
        with self.assertRaisesMessage(CommandError, "son obligatorios"):
            call_command("seed_initial_data", verbosity=0)


class ModeloBaseTests(TestCase):
    def test_alumno_belongs_to_sucursal_and_has_estado(self):
        sucursal = Sucursal.objects.create(codigo="POS", nombre="Posadas")
        alumno = Alumno.objects.create(
            legajo="A-001",
            nombre="Ana",
            apellido="Gomez",
            dni="30111222",
            email="ana@example.com",
            telefono="3764000000",
            sucursal=sucursal,
        )

        self.assertEqual(str(alumno), "Gomez, Ana")
        self.assertEqual(alumno.estado, Alumno.Estado.ACTIVO)

    def test_concepto_can_be_limited_to_carrera(self):
        sucursal = Sucursal.objects.create(codigo="ELD", nombre="Eldorado")
        carrera = CarreraCurso.objects.create(nombre="Auxiliar administrativo", sucursal=sucursal)
        concepto = ConceptoCobrable.objects.create(
            nombre="Matricula 2026",
            tipo=ConceptoCobrable.Tipo.MATRICULA,
            importe=50000,
            sucursal=sucursal,
            carrera=carrera,
        )

        self.assertEqual(str(concepto), "Matricula 2026")
        self.assertEqual(concepto.carrera, carrera)


class ImportacionDatosTests(TestCase):
    def setUp(self):
        self.posadas = Sucursal.objects.create(codigo="POS", nombre="Posadas")

    def test_helpers_normalize_names_and_mixed_dates(self):
        self.assertEqual(split_full_name("RUIZ DIAZ Aldana Micaela"), ("RUIZ DIAZ", "Aldana Micaela"))
        self.assertEqual(parse_date_value("17 de enero de 1990").isoformat(), "1990-01-17")
        self.assertEqual(parse_date_value("9/15/2001").isoformat(), "2001-09-15")

    def test_imports_canonical_templates_idempotently(self):
        class FakeReader:
            def read(self, source, filename):
                return {
                    "Carreras": [
                        ["sucursal_codigo", "nombre", "tipo", "duracion", "plan_cuotas", "importe_matricula", "cuota_total"],
                        ["POS", "Tecnicatura de prueba", "carrera", "3 años", "10", "41400", "82000"],
                    ],
                    "Alumnos": [
                        ["sucursal_codigo", "legajo", "apellido", "nombre", "dni", "cuil", "fecha_nacimiento", "email", "telefono", "domicilio", "carrera"],
                        ["POS", "POS-TEST-001", "PEREZ", "Ana", "30111222", "27301112220", "1990-01-17", "ana@example.com", "3764000000", "Centro", "Tecnicatura de prueba"],
                    ],
                }

        service = IPACWorkbookImporter(reader=FakeReader())
        first = service.import_file(b"", "plantilla.xlsx", default_branch_code="POS")
        second = service.import_file(b"", "plantilla.xlsx", default_branch_code="POS")

        self.assertEqual(first.careers.created, 1)
        self.assertEqual(first.students.created, 1)
        self.assertEqual(second.careers.updated, 1)
        self.assertEqual(second.students.updated, 1)
        self.assertEqual(Alumno.objects.get(dni="30111222").fecha_nacimiento.isoformat(), "1990-01-17")
        self.assertEqual(ConceptoCobrable.objects.get(nombre="Cuota mensual 2026").importe, 82000)

    def test_preview_uses_import_rules_without_persisting_new_records(self):
        class FakeReader:
            def read(self, source, filename):
                return {
                    "Alumnos": [
                        ["sucursal_codigo", "legajo", "apellido", "nombre", "dni"],
                        ["POS", "POS-PREVIEW-001", "PEREZ", "Ana", "30999888"],
                    ]
                }

        service = IPACWorkbookImporter(reader=FakeReader())
        before = Alumno.objects.count()

        result = service.preview_file(b"", "alumnos.csv", default_branch_code="POS")

        self.assertEqual(result.students.found, 1)
        self.assertEqual(result.students.created, 1)
        self.assertEqual(result.students.updated, 0)
        self.assertEqual(Alumno.objects.count(), before)
        self.assertFalse(Alumno.objects.filter(dni="30999888").exists())

    def test_preview_detects_updates_without_persisting_changes(self):
        Alumno.objects.create(
            legajo="POS-PREVIEW-002",
            nombre="Ana",
            apellido="PEREZ",
            dni="30999889",
            email="anterior@example.com",
            sucursal=self.posadas,
        )

        class FakeReader:
            def read(self, source, filename):
                return {
                    "Alumnos": [
                        ["sucursal_codigo", "legajo", "apellido", "nombre", "dni", "email"],
                        ["POS", "POS-PREVIEW-002", "PEREZ", "Ana", "30999889", "nuevo@example.com"],
                    ]
                }

        result = IPACWorkbookImporter(reader=FakeReader()).preview_file(b"", "alumnos.csv", default_branch_code="POS")

        self.assertEqual(result.students.created, 0)
        self.assertEqual(result.students.updated, 1)
        self.assertEqual(Alumno.objects.get(dni="30999889").email, "anterior@example.com")

    def test_preview_detects_critical_errors_without_persisting_changes(self):
        class FakeReader:
            def read(self, source, filename):
                return {
                    "Alumnos": [
                        ["sucursal_codigo", "legajo", "apellido", "nombre", "dni"],
                        ["NO_EXISTE", "PREVIEW-ERROR", "PEREZ", "Ana", "30999890"],
                    ]
                }

        before = Alumno.objects.count()
        result = IPACWorkbookImporter(reader=FakeReader()).preview_file(b"", "alumnos.csv", default_branch_code="POS")

        self.assertEqual(result.students.skipped, 1)
        self.assertGreaterEqual(len(result.errors), 1)
        self.assertEqual(Alumno.objects.count(), before)
        self.assertFalse(Alumno.objects.filter(dni="30999890").exists())

    def test_imports_concepts_and_opening_balances_idempotently_after_preview(self):
        alumno = Alumno.objects.create(
            legajo="POS-SALDO-001", nombre="Ana", apellido="PEREZ",
            dni="30999891", sucursal=self.posadas,
        )

        class FakeReader:
            def read(self, source, filename):
                return {
                    "Conceptos": [
                        ["sucursal_codigo", "nombre", "tipo", "importe", "carrera"],
                        ["POS", "Material inicial", "material", "1500", ""],
                    ],
                    "Saldos iniciales": [
                        ["sucursal_codigo", "legajo", "dni", "tipo", "importe", "fecha"],
                        ["POS", alumno.legajo, "", "deuda", "10000", "01-08-2026"],
                        ["POS", alumno.legajo, "", "saldo a favor", "2500", "01-08-2026"],
                    ],
                }

        service = IPACWorkbookImporter(reader=FakeReader())
        preview = service.preview_file(b"", "iniciales.xlsx", default_branch_code="POS")
        self.assertEqual(preview.concepts.created, 1)
        self.assertEqual(preview.opening_balances.created, 2)
        self.assertFalse(ConceptoCobrable.objects.filter(nombre="Material inicial").exists())
        self.assertEqual(Cuota.objects.count(), 0)
        self.assertEqual(Pago.objects.count(), 0)

        first = service.import_file(b"", "iniciales.xlsx", default_branch_code="POS")
        second = service.import_file(b"", "iniciales.xlsx", default_branch_code="POS")

        self.assertEqual(first.concepts.created, 1)
        self.assertEqual(first.opening_balances.created, 2)
        self.assertEqual(second.concepts.updated, 1)
        self.assertEqual(second.opening_balances.updated, 2)
        self.assertEqual(Cuota.objects.get(alumno=alumno).importe, Decimal("10000"))
        self.assertEqual(Pago.objects.get(alumno=alumno).saldo_a_favor, Decimal("2500"))
        self.assertEqual(MovimientoCaja.objects.count(), 0)


class ApiInicialTests(APITestCase):
    def setUp(self):
        self.posadas = Sucursal.objects.create(codigo="POS", nombre="Posadas")
        self.eldorado = Sucursal.objects.create(codigo="ELD", nombre="Eldorado")
        self.admin = User.objects.create_user("admin", password="admin123")
        PerfilUsuario.objects.create(
            user=self.admin,
            rol=PerfilUsuario.Rol.ADMINISTRACION,
            sucursal=self.posadas,
            puede_ver_todas_las_sucursales=True,
        )
        self.superadmin = User.objects.create_user("superadmin", password="superadmin123")
        PerfilUsuario.objects.create(
            user=self.superadmin,
            rol=PerfilUsuario.Rol.SUPERADMIN,
            sucursal=self.posadas,
            puede_ver_todas_las_sucursales=True,
        )
        self.cajero = User.objects.create_user("cajero", password="cajero123")
        PerfilUsuario.objects.create(
            user=self.cajero,
            rol=PerfilUsuario.Rol.CAJA,
            sucursal=self.posadas,
        )
        self.tesoreria = User.objects.create_user("tesoreria", password="tesoreria123")
        PerfilUsuario.objects.create(
            user=self.tesoreria,
            rol=PerfilUsuario.Rol.TESORERIA,
            sucursal=self.posadas,
        )
        self.consulta = User.objects.create_user("consulta", password="consulta123")
        PerfilUsuario.objects.create(
            user=self.consulta,
            rol=PerfilUsuario.Rol.CONSULTA,
            sucursal=self.posadas,
        )
        Alumno.objects.create(
            legajo="P-001",
            nombre="Pedro",
            apellido="Perez",
            dni="22111222",
            sucursal=self.posadas,
        )
        Alumno.objects.create(
            legajo="E-001",
            nombre="Elena",
            apellido="Silva",
            dni="23111222",
            sucursal=self.eldorado,
        )
        CarreraCurso.objects.create(nombre="Secretariado", sucursal=self.posadas)
        ConceptoCobrable.objects.create(
            nombre="Cuota mensual",
            tipo=ConceptoCobrable.Tipo.CUOTA,
            importe=25000,
            sucursal=self.posadas,
        )

    def test_login_and_current_user(self):
        client = APIClient()

        response = client.post("/api/auth/login/", {"username": "admin", "password": "admin123"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("key", response.data)

        client.credentials(HTTP_AUTHORIZATION=f"Token {response.data['key']}")
        me = client.get("/api/auth/me/")

        self.assertEqual(me.status_code, status.HTTP_200_OK)
        self.assertEqual(me.data["username"], "admin")
        self.assertEqual(me.data["perfil"]["rol"], PerfilUsuario.Rol.ADMINISTRACION)

    def test_admin_can_download_import_template(self):
        self.client.force_authenticate(self.admin)

        response = self.client.get("/api/importaciones/plantillas/alumnos/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("sucursal_codigo;legajo;apellido", response.content.decode("utf-8-sig"))

    def test_admin_can_import_csv_template(self):
        self.client.force_authenticate(self.admin)
        content = (
            "sucursal_codigo;legajo;apellido;nombre;dni;cuil;fecha_nacimiento;email;telefono;domicilio;carrera\n"
            "POS;P-CSV-001;Lopez;Ana;30999888;;1990-01-17;ana.csv@example.com;3764000011;Centro;\n"
        ).encode("utf-8")

        preview = self.client.post(
            "/api/importaciones/workbook/preview/",
            {"archivo": SimpleUploadedFile("alumnos.csv", content, content_type="text/csv"), "sucursal": "POS"},
            format="multipart",
        )
        response = self.client.post(
            "/api/importaciones/workbook/",
            {
                "archivo": SimpleUploadedFile("alumnos.csv", content, content_type="text/csv"),
                "sucursal": "POS",
                "preview_token": preview.data["preview_token"],
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["alumnos"]["created"], 1)
        self.assertTrue(Alumno.objects.filter(legajo="P-CSV-001", dni="30999888").exists())

    def test_import_requires_preview_of_the_exact_same_file(self):
        self.client.force_authenticate(self.admin)
        content = (
            "sucursal_codigo;legajo;apellido;nombre;dni\n"
            "POS;P-SIGNED-001;Lopez;Ana;30999885\n"
        ).encode("utf-8")
        direct = self.client.post(
            "/api/importaciones/workbook/",
            {"archivo": SimpleUploadedFile("signed.csv", content, content_type="text/csv"), "sucursal": "POS"},
            format="multipart",
        )
        preview = self.client.post(
            "/api/importaciones/workbook/preview/",
            {"archivo": SimpleUploadedFile("signed.csv", content, content_type="text/csv"), "sucursal": "POS"},
            format="multipart",
        )
        changed = self.client.post(
            "/api/importaciones/workbook/",
            {
                "archivo": SimpleUploadedFile("signed.csv", content + b"\n", content_type="text/csv"),
                "sucursal": "POS",
                "preview_token": preview.data["preview_token"],
            },
            format="multipart",
        )

        self.assertEqual(direct.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(changed.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(Alumno.objects.filter(legajo="P-SIGNED-001").exists())

    def test_admin_can_preview_new_student_without_modifying_database(self):
        self.client.force_authenticate(self.admin)
        content = (
            "sucursal_codigo;legajo;apellido;nombre;dni;email\n"
            "POS;P-PREVIEW-001;Lopez;Ana;30999887;preview@example.com\n"
        ).encode("utf-8")
        before = Alumno.objects.count()

        response = self.client.post(
            "/api/importaciones/workbook/preview/",
            {"archivo": SimpleUploadedFile("preview.csv", content, content_type="text/csv"), "sucursal": "POS"},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["alumnos"]["found"], 1)
        self.assertEqual(response.data["alumnos"]["created"], 1)
        self.assertEqual(response.data["total_errores"], 0)
        self.assertEqual(Alumno.objects.count(), before)

    def test_admin_preview_detects_update_without_persisting_it(self):
        self.client.force_authenticate(self.admin)
        existing = Alumno.objects.get(legajo="P-001")
        content = (
            "sucursal_codigo;legajo;apellido;nombre;dni;email\n"
            f"POS;P-001;Perez;Pedro;{existing.dni};nuevo-preview@example.com\n"
        ).encode("utf-8")

        response = self.client.post(
            "/api/importaciones/workbook/preview/",
            {"archivo": SimpleUploadedFile("preview-update.csv", content, content_type="text/csv"), "sucursal": "POS"},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["alumnos"]["updated"], 1)
        self.assertNotEqual(Alumno.objects.get(pk=existing.pk).email, "nuevo-preview@example.com")

    def test_admin_preview_reports_critical_errors_without_persisting_them(self):
        self.client.force_authenticate(self.admin)
        content = (
            "sucursal_codigo;legajo;apellido;nombre;dni\n"
            "NO_EXISTE;P-ERROR;Lopez;Ana;30999886\n"
        ).encode("utf-8")
        before = Alumno.objects.count()

        response = self.client.post(
            "/api/importaciones/workbook/preview/",
            {"archivo": SimpleUploadedFile("preview-error.csv", content, content_type="text/csv"), "sucursal": "POS"},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(response.data["total_errores"], 1)
        self.assertEqual(Alumno.objects.count(), before)
        self.assertFalse(Alumno.objects.filter(dni="30999886").exists())

    def test_non_admin_cannot_import_data(self):
        self.client.force_authenticate(self.cajero)

        response = self.client.post("/api/importaciones/workbook/", {}, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_unauthenticated_api_requests_are_rejected(self):
        response = self.client.get("/api/alumnos/")

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_consulta_can_read_but_cannot_mutate_core_resources(self):
        self.client.force_authenticate(user=self.consulta)
        today = timezone.localdate().isoformat()

        self.assertEqual(self.client.get("/api/alumnos/").status_code, status.HTTP_200_OK)
        self.assertEqual(self.client.get("/api/conceptos/").status_code, status.HTTP_200_OK)
        self.assertEqual(
            self.client.get(f"/api/reportes/resumen/?desde={today}&hasta={today}").status_code,
            status.HTTP_200_OK,
        )
        self.assertEqual(
            self.client.post(
                "/api/alumnos/",
                {"legajo": "CONSULTA-001", "nombre": "No", "apellido": "Debe", "dni": "40000001", "sucursal": self.posadas.id},
                format="json",
            ).status_code,
            status.HTTP_403_FORBIDDEN,
        )
        self.assertEqual(
            self.client.post(
                "/api/pagos/",
                {"alumno": Alumno.objects.get(legajo="P-001").id, "importe": "1000.00", "medio": Pago.Medio.EFECTIVO},
                format="json",
            ).status_code,
            status.HTTP_403_FORBIDDEN,
        )
        self.assertEqual(self.client.post("/api/movimientos-caja/", {}, format="json").status_code, status.HTTP_403_FORBIDDEN)

    def test_tesoreria_can_collect_manage_fees_cash_and_reports_but_not_users(self):
        self.client.force_authenticate(user=self.tesoreria)
        alumno = Alumno.objects.get(legajo="P-001")
        carrera = CarreraCurso.objects.get(nombre="Secretariado")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        today = timezone.localdate()
        matricula = Matricula.objects.create(alumno=alumno, carrera=carrera, sucursal=self.posadas, fecha_inicio=today)

        pago = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "importe": "1000.00", "medio": Pago.Medio.EFECTIVO},
            format="json",
        )
        cuota = self.client.post(
            "/api/cuotas/",
            {"alumno": alumno.id, "matricula": matricula.id, "concepto": concepto.id, "periodo": "2026-08", "fecha_emision": today, "fecha_vencimiento": today, "importe": "25000.00"},
            format="json",
        )

        self.assertEqual(pago.status_code, status.HTTP_201_CREATED)
        self.assertEqual(cuota.status_code, status.HTTP_201_CREATED)
        self.assertEqual(self.client.get(f"/api/cajas/hoy/?sucursal={self.posadas.id}").status_code, status.HTTP_200_OK)
        self.assertEqual(self.client.get(f"/api/reportes/resumen/?desde={today}&hasta={today}").status_code, status.HTTP_200_OK)
        self.assertEqual(self.client.get("/api/usuarios/").status_code, status.HTTP_403_FORBIDDEN)

    def test_caja_can_collect_and_operate_own_cash_but_not_configure(self):
        self.client.force_authenticate(user=self.cajero)
        alumno = Alumno.objects.get(legajo="P-001")
        pago = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "importe": "1000.00", "medio": Pago.Medio.EFECTIVO},
            format="json",
        )
        caja = self.client.get(f"/api/cajas/hoy/?sucursal={self.posadas.id}")
        movimiento = self.client.post(
            "/api/movimientos-caja/",
            {"caja": caja.data["id"], "tipo": MovimientoCaja.Tipo.EGRESO, "medio": Pago.Medio.EFECTIVO, "importe": "100.00", "descripcion": "Insumos"},
            format="json",
        )

        self.assertEqual(pago.status_code, status.HTTP_201_CREATED)
        self.assertEqual(caja.status_code, status.HTTP_200_OK)
        self.assertEqual(movimiento.status_code, status.HTTP_201_CREATED)
        self.assertEqual(self.client.post("/api/cajas/", {"sucursal": self.posadas.id}, format="json").status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(self.client.post("/api/conceptos/", {}, format="json").status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(
            self.client.post(
                "/api/cuotas/",
                {
                    "alumno": alumno.id,
                    "concepto": ConceptoCobrable.objects.get(nombre="Cuota mensual").id,
                    "periodo": "2026-12",
                    "fecha_emision": timezone.localdate(),
                    "fecha_vencimiento": timezone.localdate(),
                    "importe": "25000.00",
                },
                format="json",
            ).status_code,
            status.HTTP_403_FORBIDDEN,
        )
        self.assertEqual(
            self.client.post(
                "/api/cuotas/generar/",
                {
                    "alumnos": [alumno.id],
                    "concepto": ConceptoCobrable.objects.get(nombre="Cuota mensual").id,
                    "periodo": "2026-12",
                    "fecha_emision": timezone.localdate(),
                    "fecha_vencimiento": timezone.localdate(),
                    "importe": "25000.00",
                },
                format="json",
            ).status_code,
            status.HTTP_403_FORBIDDEN,
        )
        self.assertEqual(self.client.post("/api/usuarios/", {}, format="json").status_code, status.HTTP_403_FORBIDDEN)

    def test_administracion_cannot_escalate_or_expand_user_scope(self):
        self.client.force_authenticate(user=self.admin)
        superadmin = self.client.post(
            "/api/usuarios/",
            {"username": "new-superadmin", "password": "secret123", "rol": PerfilUsuario.Rol.SUPERADMIN, "sucursal": self.posadas.id},
            format="json",
        )
        self.assertEqual(superadmin.status_code, status.HTTP_400_BAD_REQUEST)

        target = User.objects.create_user("target", password="target123")
        PerfilUsuario.objects.create(user=target, rol=PerfilUsuario.Rol.CONSULTA, sucursal=self.posadas)
        expand_scope = self.client.patch(
            f"/api/usuarios/{target.id}/",
            {"puede_ver_todas_las_sucursales": True},
            format="json",
        )
        self.assertEqual(expand_scope.status_code, status.HTTP_400_BAD_REQUEST)

        self_update = self.client.patch(
            f"/api/usuarios/{self.admin.id}/",
            {"rol": PerfilUsuario.Rol.CONSULTA},
            format="json",
        )
        self.assertEqual(self_update.status_code, status.HTTP_400_BAD_REQUEST)

    def test_superadmin_can_manage_privileged_users(self):
        self.client.force_authenticate(user=self.superadmin)
        response = self.client.post(
            "/api/usuarios/",
            {"username": "new-superadmin", "password": "secret123", "rol": PerfilUsuario.Rol.SUPERADMIN, "sucursal": self.posadas.id, "puede_ver_todas_las_sucursales": True},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["perfil"]["rol"], PerfilUsuario.Rol.SUPERADMIN)

    def test_user_without_global_access_only_sees_own_sucursal_students(self):
        self.client.force_authenticate(user=self.cajero)

        response = self.client.get("/api/alumnos/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual([item["legajo"] for item in response.data["results"]], ["P-001"])

    def test_admin_can_list_initial_catalogs(self):
        self.client.force_authenticate(user=self.admin)

        sucursales = self.client.get("/api/sucursales/")
        carreras = self.client.get("/api/carreras/")
        conceptos = self.client.get("/api/conceptos/")

        self.assertEqual(sucursales.status_code, status.HTTP_200_OK)
        self.assertEqual(carreras.status_code, status.HTTP_200_OK)
        self.assertEqual(conceptos.status_code, status.HTTP_200_OK)
        self.assertEqual(sucursales.data["count"], 2)
        self.assertEqual(carreras.data["count"], 1)
        self.assertEqual(conceptos.data["count"], 1)

    def test_authenticated_user_can_update_student(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")

        response = self.client.patch(
            f"/api/alumnos/{alumno.id}/",
            {"telefono": "3764555010", "email": "pedro.actualizado@example.com"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        alumno.refresh_from_db()
        self.assertEqual(alumno.telefono, "3764555010")
        self.assertEqual(alumno.email, "pedro.actualizado@example.com")

    def test_authenticated_user_can_register_payment_for_student(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual", sucursal=self.posadas)

        response = self.client.post(
            "/api/pagos/",
            {
                "alumno": alumno.id,
                "concepto": concepto.id,
                "importe": "25000.00",
                "medio": Pago.Medio.EFECTIVO,
                "observacion": "Pago de prueba",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Pago.objects.count(), 1)
        self.assertEqual(response.data["alumno_nombre"], "Perez, Pedro")

    def test_payment_list_can_be_filtered_by_student(self):
        self.client.force_authenticate(user=self.admin)
        pedro = Alumno.objects.get(legajo="P-001")
        elena = Alumno.objects.get(legajo="E-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual", sucursal=self.posadas)
        Pago.objects.create(
            alumno=pedro,
            concepto=concepto,
            sucursal=self.posadas,
            importe=25000,
            medio=Pago.Medio.EFECTIVO,
        )
        Pago.objects.create(
            alumno=elena,
            sucursal=self.eldorado,
            importe=12000,
            medio=Pago.Medio.TRANSFERENCIA,
        )

        response = self.client.get(f"/api/pagos/?alumno={pedro.id}")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["alumno"], pedro.id)

    def test_student_search_finds_records_after_first_page_by_all_identifiers(self):
        self.client.force_authenticate(user=self.admin)
        for index in range(55):
            Alumno.objects.create(
                legajo=f"Z-{index:03d}",
                nombre=f"Nombre {index}",
                apellido=f"Zeta {index:03d}",
                dni=f"40{index:06d}",
                sucursal=self.posadas,
            )
        target = Alumno.objects.create(
            legajo="TARGET-999",
            nombre="Lucia",
            apellido="Zzz",
            dni="49999999",
            sucursal=self.posadas,
        )

        second_page = self.client.get("/api/alumnos/?page=3&page_size=25")
        self.assertEqual(second_page.status_code, status.HTTP_200_OK)
        self.assertEqual(second_page.data["count"], 58)
        self.assertEqual(second_page.data["page"], 3)
        self.assertEqual(second_page.data["page_size"], 25)
        self.assertIn(target.id, [item["id"] for item in second_page.data["results"]])

        for term in (target.nombre, target.apellido, target.dni, target.legajo):
            response = self.client.get("/api/alumnos/", {"search": term})
            self.assertEqual(response.status_code, status.HTTP_200_OK)
            self.assertEqual(response.data["count"], 1)
            self.assertEqual(response.data["results"][0]["id"], target.id)

    def test_student_filters_apply_server_side_by_branch_state_and_career(self):
        self.client.force_authenticate(user=self.admin)
        carrera = CarreraCurso.objects.create(nombre="Curso filtrable", sucursal=self.eldorado)
        target = Alumno.objects.create(
            legajo="FILTER-001",
            nombre="Alumno",
            apellido="Filtrado",
            dni="48888888",
            estado=Alumno.Estado.BAJA,
            sucursal=self.eldorado,
            carrera=carrera,
        )

        response = self.client.get(
            "/api/alumnos/",
            {"sucursal": self.eldorado.id, "estado": Alumno.Estado.BAJA, "carrera": carrera.id},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["id"], target.id)

    def test_student_filters_include_backend_financial_status(self):
        self.client.force_authenticate(user=self.admin)
        debtor = Alumno.objects.get(legajo="P-001")
        creditor = Alumno.objects.create(
            legajo="CREDIT-001",
            nombre="Saldo",
            apellido="Favor",
            dni="47777666",
            sucursal=self.posadas,
        )
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        Cuota.objects.create(
            alumno=debtor,
            concepto=concepto,
            sucursal=self.posadas,
            periodo="08-2026",
            fecha_emision=timezone.localdate(),
            fecha_vencimiento=timezone.localdate(),
            importe="15000.00",
        )
        Pago.objects.create(
            alumno=creditor,
            sucursal=self.posadas,
            importe="9000.00",
            medio=Pago.Medio.TRANSFERENCIA,
        )

        debts = self.client.get("/api/alumnos/", {"con_deuda": "1"})
        credits = self.client.get("/api/alumnos/", {"con_saldo_favor": "1"})

        self.assertEqual(debts.status_code, status.HTTP_200_OK)
        self.assertEqual(debts.data["count"], 1)
        self.assertEqual(debts.data["results"][0]["id"], debtor.id)
        self.assertEqual(debts.data["results"][0]["deuda_total"], "15000.00")
        self.assertEqual(credits.data["count"], 1)
        self.assertEqual(credits.data["results"][0]["id"], creditor.id)
        self.assertEqual(credits.data["results"][0]["saldo_a_favor"], "9000.00")

    def test_student_statistics_count_all_matching_records_not_current_page(self):
        self.client.force_authenticate(user=self.admin)
        for index in range(55):
            Alumno.objects.create(
                legajo=f"STAT-{index:03d}",
                nombre=f"Activo {index}",
                apellido="Estadistica",
                dni=f"470{index:05d}",
                sucursal=self.posadas,
            )
        for index in range(3):
            Alumno.objects.create(
                legajo=f"STAT-I-{index:03d}",
                nombre=f"Inactivo {index}",
                apellido="Estadistica",
                dni=f"480{index:05d}",
                estado=Alumno.Estado.BAJA,
                sucursal=self.posadas,
            )

        response = self.client.get(f"/api/alumnos/estadisticas/?sucursal={self.posadas.id}")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"total": 59, "activos": 56, "inactivos": 3})

    def test_get_today_cashbox_creates_open_cashbox(self):
        self.client.force_authenticate(user=self.admin)

        response = self.client.get(f"/api/cajas/hoy/?sucursal={self.posadas.id}")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["estado"], CajaDiaria.Estado.ABIERTA)
        self.assertEqual(response.data["sucursal"], self.posadas.id)
        self.assertEqual(CajaDiaria.objects.count(), 1)

    def test_payment_creates_cash_movement(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual", sucursal=self.posadas)

        response = self.client.post(
            "/api/pagos/",
            {
                "alumno": alumno.id,
                "concepto": concepto.id,
                "importe": "25000.00",
                "medio": Pago.Medio.EFECTIVO,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        movimiento = MovimientoCaja.objects.get(pago_id=response.data["id"])
        self.assertEqual(movimiento.tipo, MovimientoCaja.Tipo.PAGO)
        self.assertEqual(movimiento.importe, Pago.objects.get(id=response.data["id"]).importe)
        self.assertEqual(movimiento.caja.sucursal, self.posadas)

    def test_payment_creates_open_cashbox_when_daily_cashbox_does_not_exist(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")

        response = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "importe": "1500.00", "medio": Pago.Medio.EFECTIVO},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        caja = CajaDiaria.objects.get(fecha=timezone.localdate(), sucursal=self.posadas, usuario=self.admin)
        self.assertEqual(caja.estado, CajaDiaria.Estado.ABIERTA)
        self.assertEqual(str(caja.total_esperado), "1500.00")

    def test_payment_updates_an_existing_open_cashbox(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        caja = CajaDiaria.objects.create(
            fecha=timezone.localdate(),
            sucursal=self.posadas,
            usuario=self.admin,
            estado=CajaDiaria.Estado.ABIERTA,
        )

        response = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "importe": "2200.00", "medio": Pago.Medio.TRANSFERENCIA},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        caja.refresh_from_db()
        self.assertEqual(str(caja.total_esperado), "0")
        self.assertEqual(str(caja.diferencia), "0.00")
        self.assertEqual(str(caja.resumen.total_cobrado), "2200.00")
        self.assertEqual(str(caja.resumen.transferencia), "2200.00")

    def test_cashbox_summary_separates_physical_cash_from_electronic_payments(self):
        caja = CajaDiaria.objects.create(
            fecha=timezone.localdate(),
            sucursal=self.posadas,
            usuario=self.cajero,
            saldo_inicial=Decimal("20000.00"),
        )
        MovimientoCaja.objects.create(
            caja=caja,
            tipo=MovimientoCaja.Tipo.PAGO,
            medio=Pago.Medio.EFECTIVO,
            importe=Decimal("150000.00"),
        )
        MovimientoCaja.objects.create(
            caja=caja,
            tipo=MovimientoCaja.Tipo.PAGO,
            medio=Pago.Medio.TRANSFERENCIA,
            importe=Decimal("80000.00"),
        )
        MovimientoCaja.objects.create(
            caja=caja,
            tipo=MovimientoCaja.Tipo.PAGO,
            medio=Pago.Medio.MERCADO_PAGO,
            importe=Decimal("40000.00"),
        )
        MovimientoCaja.objects.create(
            caja=caja,
            tipo=MovimientoCaja.Tipo.EGRESO,
            medio=Pago.Medio.EFECTIVO,
            importe=Decimal("10000.00"),
        )

        resumen = caja.resumen

        self.assertEqual(resumen.efectivo_esperado, Decimal("160000.00"))
        self.assertEqual(resumen.total_cobrado, Decimal("270000.00"))
        self.assertEqual(resumen.transferencia, Decimal("80000.00"))
        self.assertEqual(resumen.mercado_pago, Decimal("40000.00"))

    def test_partial_carry_can_be_consumed_once_by_another_user_after_days_without_operation(self):
        self.client.force_authenticate(user=self.admin)
        origen = CajaDiaria.objects.create(
            fecha=timezone.localdate() - timedelta(days=3),
            sucursal=self.posadas,
            usuario=self.admin,
        )
        MovimientoCaja.objects.create(
            caja=origen,
            tipo=MovimientoCaja.Tipo.INGRESO,
            medio=Pago.Medio.EFECTIVO,
            importe=Decimal("160000.00"),
        )
        cierre = self.client.post(
            f"/api/cajas/{origen.id}/cerrar/",
            {
                "total_contado": "160000.00",
                "importe_retirado": "110000.00",
                "saldo_arrastrable": "50000.00",
            },
            format="json",
        )
        self.assertEqual(cierre.status_code, status.HTTP_200_OK)
        self.assertEqual(cierre.data["saldo_arrastrable"], "50000.00")

        self.client.force_authenticate(user=self.cajero)
        destino = self.client.get(f"/api/cajas/hoy/?sucursal={self.posadas.id}").data
        disponible = self.client.get(f"/api/cajas/{destino['id']}/saldo-anterior/")
        self.assertEqual(disponible.status_code, status.HTTP_200_OK)
        self.assertTrue(disponible.data["disponible"])
        self.assertEqual(disponible.data["importe"], Decimal("50000.00"))
        self.assertEqual(disponible.data["usuario_origen"], self.admin.username)

        aplicado = self.client.post(
            f"/api/cajas/{destino['id']}/saldo-anterior/",
            {"saldo_id": disponible.data["id"]},
            format="json",
        )
        self.assertEqual(aplicado.status_code, status.HTTP_200_OK)
        self.assertEqual(aplicado.data["saldo_inicial"], "50000.00")
        self.assertEqual(aplicado.data["total_esperado"], "50000.00")

        saldo = SaldoArrastrableCaja.objects.get(caja_origen=origen)
        self.assertEqual(saldo.caja_destino_id, destino["id"])
        self.assertIsNotNone(saldo.utilizado_en)
        segundo_intento = self.client.post(
            f"/api/cajas/{destino['id']}/saldo-anterior/",
            {"saldo_id": saldo.id},
            format="json",
        )
        self.assertEqual(segundo_intento.status_code, status.HTTP_400_BAD_REQUEST)

    def test_close_can_withdraw_all_cash_without_creating_carry(self):
        self.client.force_authenticate(user=self.admin)
        caja = CajaDiaria.objects.create(
            fecha=timezone.localdate(),
            sucursal=self.posadas,
            usuario=self.admin,
        )

        cierre = self.client.post(
            f"/api/cajas/{caja.id}/cerrar/",
            {
                "total_contado": "12000.00",
                "importe_retirado": "12000.00",
                "saldo_arrastrable": "0.00",
            },
            format="json",
        )

        self.assertEqual(cierre.status_code, status.HTTP_200_OK)
        self.assertFalse(SaldoArrastrableCaja.objects.filter(caja_origen=caja).exists())

    def test_close_rejects_carry_above_counted_cash_without_partial_effects(self):
        self.client.force_authenticate(user=self.admin)
        caja = CajaDiaria.objects.create(
            fecha=timezone.localdate(),
            sucursal=self.posadas,
            usuario=self.admin,
        )

        cierre = self.client.post(
            f"/api/cajas/{caja.id}/cerrar/",
            {
                "total_contado": "10000.00",
                "importe_retirado": "0.00",
                "saldo_arrastrable": "12000.00",
            },
            format="json",
        )

        self.assertEqual(cierre.status_code, status.HTTP_400_BAD_REQUEST)
        caja.refresh_from_db()
        self.assertEqual(caja.estado, CajaDiaria.Estado.ABIERTA)
        self.assertFalse(SaldoArrastrableCaja.objects.filter(caja_origen=caja).exists())

    def test_payment_is_rejected_without_side_effects_when_cashbox_is_closed(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        caja = CajaDiaria.objects.create(
            fecha=timezone.localdate(),
            sucursal=self.posadas,
            usuario=self.admin,
            estado=CajaDiaria.Estado.CERRADA,
            total_contado=Decimal("3500.00"),
            cerrada_en=timezone.now(),
        )
        expected_before = caja.total_esperado
        difference_before = caja.diferencia
        payments_before = Pago.objects.count()
        movements_before = MovimientoCaja.objects.count()

        response = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "importe": "2200.00", "medio": Pago.Medio.EFECTIVO},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("La caja del día está cerrada. No se pueden registrar nuevas cobranzas.", str(response.data))
        self.assertEqual(Pago.objects.count(), payments_before)
        self.assertEqual(MovimientoCaja.objects.count(), movements_before)
        caja.refresh_from_db()
        self.assertEqual(caja.total_esperado, expected_before)
        self.assertEqual(caja.diferencia, difference_before)

    def test_manual_movement_keeps_closed_cashbox_immutable(self):
        self.client.force_authenticate(user=self.admin)
        caja = CajaDiaria.objects.create(
            fecha=timezone.localdate(),
            sucursal=self.posadas,
            usuario=self.admin,
            estado=CajaDiaria.Estado.CERRADA,
            total_contado=Decimal("3500.00"),
            cerrada_en=timezone.now(),
        )
        expected_before = caja.total_esperado
        difference_before = caja.diferencia

        response = self.client.post(
            "/api/movimientos-caja/",
            {"caja": caja.id, "tipo": MovimientoCaja.Tipo.INGRESO, "medio": Pago.Medio.EFECTIVO, "importe": "900.00", "descripcion": "Intento posterior al cierre"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("La caja del día está cerrada. No se pueden registrar nuevas cobranzas.", str(response.data))
        self.assertEqual(MovimientoCaja.objects.filter(caja=caja).count(), 0)
        caja.refresh_from_db()
        self.assertEqual(caja.total_esperado, expected_before)
        self.assertEqual(caja.diferencia, difference_before)

    def test_can_register_manual_cashbox_movement_and_close_cashbox(self):
        self.client.force_authenticate(user=self.admin)
        caja = self.client.get(f"/api/cajas/hoy/?sucursal={self.posadas.id}").data

        movimiento = self.client.post(
            "/api/movimientos-caja/",
            {
                "caja": caja["id"],
                "tipo": MovimientoCaja.Tipo.EGRESO,
                "medio": Pago.Medio.EFECTIVO,
                "importe": "5000.00",
                "descripcion": "Compra de insumos",
            },
            format="json",
        )
        self.assertEqual(movimiento.status_code, status.HTTP_201_CREATED)

        cierre = self.client.post(
            f"/api/cajas/{caja['id']}/cerrar/",
            {"total_contado": "10000.00"},
            format="json",
        )

        self.assertEqual(cierre.status_code, status.HTTP_200_OK)
        self.assertEqual(cierre.data["estado"], CajaDiaria.Estado.CERRADA)
        self.assertEqual(cierre.data["total_contado"], "10000.00")
        self.assertEqual(cierre.data["total_esperado"], "-5000.00")
        self.assertEqual(cierre.data["diferencia"], "15000.00")

    def test_branch_user_cannot_register_payment_for_other_branch_student(self):
        self.client.force_authenticate(user=self.cajero)
        elena = Alumno.objects.get(legajo="E-001")

        response = self.client.post(
            "/api/pagos/",
            {
                "alumno": elena.id,
                "importe": "12000.00",
                "medio": Pago.Medio.EFECTIVO,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Pago.objects.count(), 0)

    def test_branch_user_cannot_register_movement_in_another_users_cashbox(self):
        self.client.force_authenticate(user=self.admin)
        admin_cashbox = self.client.get(f"/api/cajas/hoy/?sucursal={self.posadas.id}").data

        self.client.force_authenticate(user=self.cajero)
        response = self.client.post(
            "/api/movimientos-caja/",
            {
                "caja": admin_cashbox["id"],
                "tipo": MovimientoCaja.Tipo.EGRESO,
                "medio": Pago.Medio.EFECTIVO,
                "importe": "1000.00",
                "descripcion": "Intento no permitido",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(MovimientoCaja.objects.count(), 0)

    def test_can_enroll_student_and_create_fee(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        carrera = CarreraCurso.objects.get(nombre="Secretariado")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        hoy = timezone.localdate()

        matricula = self.client.post("/api/matriculas/", {"alumno": alumno.id, "carrera": carrera.id, "fecha_inicio": hoy}, format="json")
        self.assertEqual(matricula.status_code, status.HTTP_201_CREATED)

        cuota = self.client.post("/api/cuotas/", {"alumno": alumno.id, "matricula": matricula.data["id"], "concepto": concepto.id, "periodo": "2026-07", "fecha_emision": hoy, "fecha_vencimiento": hoy, "importe": "25000.00", "descuento": "2000.00", "recargo": "500.00"}, format="json")
        self.assertEqual(cuota.status_code, status.HTTP_201_CREATED)
        self.assertEqual(cuota.data["total"], "23500.00")
        self.assertEqual(cuota.data["saldo"], "23500.00")

    def test_matricula_creation_updates_legacy_student_career_and_history(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        carrera = CarreraCurso.objects.get(nombre="Secretariado")
        hoy = timezone.localdate()

        response = self.client.post(
            "/api/matriculas/",
            {
                "alumno": alumno.id,
                "carrera": carrera.id,
                "fecha_inicio": hoy,
                "observacion": "Ingreso al ciclo 2026",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        alumno.refresh_from_db()
        self.assertEqual(alumno.carrera_id, carrera.id)
        history = self.client.get(f"/api/matriculas/?alumno={alumno.id}")
        self.assertEqual(history.status_code, status.HTTP_200_OK)
        self.assertEqual(history.data["count"], 1)
        self.assertEqual(history.data["results"][0]["estado"], Matricula.Estado.ACTIVA)

    def test_finalizar_matricula_preserves_history_and_clears_legacy_active_career(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        carrera = CarreraCurso.objects.get(nombre="Secretariado")
        hoy = timezone.localdate()
        created = self.client.post(
            "/api/matriculas/",
            {"alumno": alumno.id, "carrera": carrera.id, "fecha_inicio": hoy},
            format="json",
        )

        response = self.client.post(f"/api/matriculas/{created.data['id']}/finalizar/", {}, format="json")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["estado"], Matricula.Estado.FINALIZADA)
        self.assertEqual(response.data["fecha_fin"], hoy.isoformat())
        self.assertEqual(Matricula.objects.count(), 1)
        alumno.refresh_from_db()
        self.assertIsNone(alumno.carrera_id)

    def test_matricula_rejects_cross_branch_and_duplicate_active_career(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        carrera_posadas = CarreraCurso.objects.get(nombre="Secretariado")
        carrera_eldorado = CarreraCurso.objects.create(nombre="Tecnicatura Eldorado", sucursal=self.eldorado)
        hoy = timezone.localdate()

        cross_branch = self.client.post(
            "/api/matriculas/",
            {"alumno": alumno.id, "carrera": carrera_eldorado.id, "fecha_inicio": hoy},
            format="json",
        )
        first = self.client.post(
            "/api/matriculas/",
            {"alumno": alumno.id, "carrera": carrera_posadas.id, "fecha_inicio": hoy},
            format="json",
        )
        duplicate = self.client.post(
            "/api/matriculas/",
            {"alumno": alumno.id, "carrera": carrera_posadas.id, "fecha_inicio": hoy},
            format="json",
        )

        self.assertEqual(cross_branch.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(first.status_code, status.HTTP_201_CREATED)
        self.assertEqual(duplicate.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Matricula.objects.count(), 1)

    def test_change_career_finalizes_previous_enrollment_and_preserves_history(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        original = CarreraCurso.objects.get(nombre="Secretariado")
        replacement = CarreraCurso.objects.create(nombre="Administración avanzada", sucursal=self.posadas)
        today = timezone.localdate()
        created = self.client.post("/api/matriculas/", {"alumno": alumno.id, "carrera": original.id, "fecha_inicio": today}, format="json")

        changed = self.client.post(
            f"/api/matriculas/{created.data['id']}/cambiar-carrera/",
            {"carrera": replacement.id, "fecha_inicio": today, "observacion": "Cambio solicitado"},
            format="json",
        )

        self.assertEqual(changed.status_code, status.HTTP_201_CREATED)
        self.assertEqual(changed.data["carrera"], replacement.id)
        self.assertEqual(Matricula.objects.filter(alumno=alumno, estado=Matricula.Estado.ACTIVA).count(), 1)
        self.assertEqual(Matricula.objects.get(pk=created.data["id"]).estado, Matricula.Estado.FINALIZADA)
        alumno.refresh_from_db()
        self.assertEqual(alumno.carrera_id, replacement.id)

    def test_annul_enrollment_requires_reason_and_clears_active_career(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        carrera = CarreraCurso.objects.get(nombre="Secretariado")
        created = self.client.post("/api/matriculas/", {"alumno": alumno.id, "carrera": carrera.id, "fecha_inicio": timezone.localdate()}, format="json")

        missing_reason = self.client.post(f"/api/matriculas/{created.data['id']}/anular/", {}, format="json")
        annulled = self.client.post(f"/api/matriculas/{created.data['id']}/anular/", {"motivo": "Carga duplicada"}, format="json")

        self.assertEqual(missing_reason.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(annulled.status_code, status.HTTP_200_OK)
        self.assertEqual(annulled.data["estado"], Matricula.Estado.ANULADA)
        self.assertIn("Carga duplicada", annulled.data["observacion"])
        alumno.refresh_from_db()
        self.assertIsNone(alumno.carrera_id)

    def test_partial_payment_updates_fee_and_keeps_credit_balance(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        hoy = timezone.localdate()
        cuota = Cuota.objects.create(alumno=alumno, concepto=concepto, sucursal=self.posadas, periodo="2026-07", fecha_emision=hoy, fecha_vencimiento=hoy, importe=25000)

        pago = self.client.post("/api/pagos/", {"alumno": alumno.id, "importe": "30000.00", "medio": Pago.Medio.EFECTIVO}, format="json")
        aplicacion = self.client.post("/api/aplicaciones-pago/", {"pago": pago.data["id"], "cuota": cuota.id, "importe": "10000.00"}, format="json")

        self.assertEqual(aplicacion.status_code, status.HTTP_201_CREATED)
        cuota.refresh_from_db()
        self.assertEqual(cuota.estado, Cuota.Estado.PARCIAL)
        detalle_pago = self.client.get(f"/api/pagos/{pago.data['id']}/")
        self.assertEqual(detalle_pago.data["importe_aplicado"], "10000.00")
        self.assertEqual(detalle_pago.data["saldo_a_favor"], "20000.00")
        self.assertEqual(detalle_pago.data["saldo_pendiente_posterior"], "15000.00")

    def _create_fee(self, period, amount="25000.00"):
        alumno = Alumno.objects.get(legajo="P-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        hoy = timezone.localdate()
        return Cuota.objects.create(
            alumno=alumno,
            concepto=concepto,
            sucursal=self.posadas,
            periodo=period,
            fecha_emision=hoy,
            fecha_vencimiento=hoy,
            importe=amount,
        )

    def test_registering_total_payment_applies_fee_and_marks_it_paid(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        cuota = self._create_fee("2026-09")

        response = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "cuota": cuota.id, "importe": "25000.00", "medio": Pago.Medio.EFECTIVO},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        cuota.refresh_from_db()
        aplicacion = cuota.aplicaciones.get()
        self.assertEqual(aplicacion.importe, 25000)
        self.assertEqual(cuota.estado, Cuota.Estado.PAGADA)
        self.assertEqual(cuota.saldo, 0)
        self.assertTrue(MovimientoCaja.objects.filter(pago_id=response.data["id"]).exists())

    def test_registering_partial_payment_updates_fee_and_account_statement(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        cuota = self._create_fee("2026-10")

        response = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "cuota": cuota.id, "importe": "10000.00", "medio": Pago.Medio.TRANSFERENCIA},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        cuota.refresh_from_db()
        self.assertEqual(cuota.estado, Cuota.Estado.PARCIAL)
        self.assertEqual(cuota.saldo, 15000)
        estado = self.client.get(f"/api/alumnos/{alumno.id}/estado-cuenta/")
        self.assertEqual(estado.data["cuotas"][0]["saldo"], "15000.00")
        self.assertEqual(estado.data["cuotas"][0]["estado"], Cuota.Estado.PARCIAL)

    def test_payment_above_fee_balance_keeps_excess_as_credit(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        cuota = self._create_fee("2026-11")

        response = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "cuota": cuota.id, "importe": "30000.00", "medio": Pago.Medio.EFECTIVO},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        cuota.refresh_from_db()
        pago = Pago.objects.get(pk=response.data["id"])
        self.assertEqual(cuota.estado, Cuota.Estado.PAGADA)
        self.assertEqual(cuota.aplicaciones.get().importe, 25000)
        self.assertEqual(pago.saldo_a_favor, 5000)
        self.assertEqual(response.data["saldo_a_favor"], "5000.00")

    def test_payment_on_account_does_not_apply_to_a_fee(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")

        response = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "importe": "12000.00", "medio": Pago.Medio.TRANSFERENCIA},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        pago = Pago.objects.get(pk=response.data["id"])
        self.assertFalse(pago.aplicaciones.exists())
        self.assertEqual(pago.saldo_a_favor, 12000)
        self.assertTrue(MovimientoCaja.objects.filter(pago=pago).exists())

    def test_registering_payment_rolls_back_if_cash_movement_fails(self):
        alumno = Alumno.objects.get(legajo="P-001")
        cuota = self._create_fee("2026-12")
        before = (Pago.objects.count(), MovimientoCaja.objects.count(), AplicacionPago.objects.count())

        with patch(
            "core.contexts.cobranzas.application.registrar_pago.MovimientoCaja.objects.create",
            side_effect=RuntimeError("fallo de caja"),
        ):
            with self.assertRaises(RuntimeError):
                RegistrarPago().execute(
                    user=self.admin,
                    alumno=alumno,
                    importe="10000.00",
                    medio=Pago.Medio.EFECTIVO,
                    cuota=cuota,
                )

        self.assertEqual((Pago.objects.count(), MovimientoCaja.objects.count(), AplicacionPago.objects.count()), before)
        cuota.refresh_from_db()
        self.assertEqual(cuota.estado, Cuota.Estado.PENDIENTE)

    def test_payment_receipt_has_stable_number_and_details(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        cuota = self._create_fee("2027-01", amount="25000.00")

        pago = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "cuota": cuota.id, "importe": "12500.00", "medio": Pago.Medio.TRANSFERENCIA},
            format="json",
        )

        self.assertEqual(pago.status_code, status.HTTP_201_CREATED)
        self.assertTrue(pago.data["numero_recibo"].startswith("REC-"))
        recibo = self.client.get(f"/api/pagos/{pago.data['id']}/recibo/")
        self.assertEqual(recibo.status_code, status.HTTP_200_OK)
        self.assertEqual(recibo.data["numero"], pago.data["numero_recibo"])
        self.assertEqual(recibo.data["pago"]["alumno_nombre"], "Perez, Pedro")
        self.assertEqual(recibo.data["pago"]["saldo_pendiente_posterior"], "12500.00")
        self.assertEqual(recibo.data["pago"]["usuario_nombre"], self.admin.username)

    def test_automatic_payment_applies_oldest_fees_and_supports_multiple_fees(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        primera = self._create_fee("2027-02")
        segunda = self._create_fee("2027-03")

        response = self.client.post(
            "/api/pagos/",
            {
                "alumno": alumno.id,
                "importe": "30000.00",
                "medio": Pago.Medio.EFECTIVO,
                "aplicacion_automatica": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        primera.refresh_from_db()
        segunda.refresh_from_db()
        self.assertEqual(primera.estado, Cuota.Estado.PAGADA)
        self.assertEqual(segunda.estado, Cuota.Estado.PARCIAL)
        self.assertEqual(primera.saldo, Decimal("0"))
        self.assertEqual(segunda.saldo, Decimal("20000.00"))
        self.assertEqual(
            list(AplicacionPago.objects.filter(pago_id=response.data["id"]).values_list("importe", flat=True)),
            [Decimal("25000.00"), Decimal("5000.00")],
        )

    def test_manual_payment_only_applies_selected_fees(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        no_seleccionada = self._create_fee("2027-04")
        seleccionada = self._create_fee("2027-05")

        response = self.client.post(
            "/api/pagos/",
            {
                "alumno": alumno.id,
                "cuotas": [seleccionada.id],
                "importe": "10000.00",
                "medio": Pago.Medio.TRANSFERENCIA,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        no_seleccionada.refresh_from_db()
        seleccionada.refresh_from_db()
        self.assertEqual(no_seleccionada.estado, Cuota.Estado.PENDIENTE)
        self.assertEqual(seleccionada.estado, Cuota.Estado.PARCIAL)
        self.assertEqual(seleccionada.saldo, Decimal("15000.00"))

    def test_authorized_payment_void_restores_fees_and_creates_inverse_cash_movement(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        cuota = self._create_fee("2027-06")
        created = self.client.post(
            "/api/pagos/",
            {
                "alumno": alumno.id,
                "cuota": cuota.id,
                "importe": "25000.00",
                "medio": Pago.Medio.EFECTIVO,
            },
            format="json",
        )
        numero = created.data["numero_recibo"]

        self.client.force_authenticate(user=self.tesoreria)
        response = self.client.post(
            f"/api/pagos/{created.data['id']}/anular/",
            {"motivo": "Cobranza cargada por duplicado"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["estado"], Pago.Estado.ANULADO)
        self.assertEqual(response.data["numero_recibo"], numero)
        cuota.refresh_from_db()
        self.assertEqual(cuota.estado, Cuota.Estado.PENDIENTE)
        self.assertEqual(cuota.saldo, Decimal("25000.00"))
        aplicacion = AplicacionPago.objects.get(pago_id=created.data["id"])
        self.assertFalse(aplicacion.activa)
        reverso = MovimientoCaja.objects.get(movimiento_origen__pago_id=created.data["id"])
        self.assertEqual(reverso.tipo, MovimientoCaja.Tipo.REVERSO)
        self.assertEqual(reverso.caja.usuario, self.tesoreria)

        recibo = self.client.get(f"/api/pagos/{created.data['id']}/recibo/")
        self.assertEqual(recibo.data["pago"]["estado"], Pago.Estado.ANULADO)
        self.assertFalse(recibo.data["aplicaciones"][0]["activa"])

    def test_unauthorized_user_cannot_void_payment(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        created = self.client.post(
            "/api/pagos/",
            {"alumno": alumno.id, "importe": "5000.00", "medio": Pago.Medio.EFECTIVO},
            format="json",
        )

        response = self.client.post(
            f"/api/pagos/{created.data['id']}/anular/",
            {"motivo": "No autorizado"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(Pago.objects.get(pk=created.data["id"]).estado, Pago.Estado.ACTIVO)

    def test_account_statement_returns_real_fees_and_credit(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        hoy = timezone.localdate()
        cuota = Cuota.objects.create(
            alumno=alumno,
            concepto=concepto,
            sucursal=self.posadas,
            periodo="2026-08",
            fecha_emision=hoy,
            fecha_vencimiento=hoy,
            importe="25000.00",
        )
        pago = Pago.objects.create(
            alumno=alumno,
            sucursal=self.posadas,
            importe="30000.00",
            medio=Pago.Medio.TRANSFERENCIA,
        )
        aplicacion = self.client.post(
            "/api/aplicaciones-pago/",
            {"pago": pago.id, "cuota": cuota.id, "importe": "10000.00"},
            format="json",
        )
        self.assertEqual(aplicacion.status_code, status.HTTP_201_CREATED)

        estado = self.client.get(f"/api/alumnos/{alumno.id}/estado-cuenta/")
        self.assertEqual(estado.status_code, status.HTTP_200_OK)
        self.assertEqual(str(estado.data["resumen"]["saldo_pendiente"]), "15000.00")
        self.assertEqual(str(estado.data["resumen"]["saldo_a_favor"]), "20000.00")
        self.assertEqual(str(estado.data["resumen"]["saldo_neto"]), "-5000.00")

    def test_can_generate_fees_for_multiple_students_once(self):
        self.client.force_authenticate(user=self.admin)
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        pedro = Alumno.objects.get(legajo="P-001")
        ana = Alumno.objects.create(
            legajo="P-003",
            nombre="Ana",
            apellido="Lopez",
            dni="24111222",
            sucursal=self.posadas,
        )
        hoy = timezone.localdate()
        payload = {
            "alumnos": [pedro.id, ana.id],
            "concepto": concepto.id,
            "periodo": "2026-09",
            "fecha_emision": hoy,
            "fecha_vencimiento": hoy,
            "importe": "26000.00",
        }

        response = self.client.post("/api/cuotas/generar/", payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(response.data), 2)
        self.assertEqual(Cuota.objects.filter(periodo="2026-09").count(), 2)

        repetida = self.client.post("/api/cuotas/generar/", payload, format="json")
        self.assertEqual(repetida.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Cuota.objects.filter(periodo="2026-09").count(), 2)

    def test_can_evaluate_mass_fee_generation_without_loading_all_pages(self):
        self.client.force_authenticate(user=self.admin)
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        pedro = Alumno.objects.get(legajo="P-001")
        ana = Alumno.objects.create(
            legajo="P-003",
            nombre="Ana",
            apellido="Lopez",
            dni="24111222",
            sucursal=self.posadas,
        )
        hoy = timezone.localdate()
        Cuota.objects.create(
            alumno=pedro,
            concepto=concepto,
            sucursal=self.posadas,
            periodo="2026-10",
            fecha_emision=hoy,
            fecha_vencimiento=hoy,
            importe="26000.00",
        )

        response = self.client.post(
            "/api/cuotas/evaluar-generacion/",
            {"sucursal": self.posadas.id, "concepto": concepto.id, "periodo": "2026-10"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["alumnos_encontrados"], 2)
        self.assertEqual(response.data["omitidas"], 1)
        self.assertEqual(response.data["alumnos_elegibles"], [ana.id])

    def test_operational_report_respects_branch_and_date_range(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        hoy = timezone.localdate()
        Cuota.objects.create(
            alumno=alumno,
            concepto=concepto,
            sucursal=self.posadas,
            periodo="2027-02",
            fecha_emision=hoy,
            fecha_vencimiento=hoy,
            importe="20000.00",
        )
        Pago.objects.create(alumno=alumno, sucursal=self.posadas, importe="12500.00", medio=Pago.Medio.EFECTIVO)

        reporte = self.client.get(
            f"/api/reportes/resumen/?desde={timezone.localdate()}&hasta={timezone.localdate()}&sucursal={self.posadas.id}"
        )
        self.assertEqual(reporte.status_code, status.HTTP_200_OK)
        self.assertEqual(reporte.data["cobranzas"]["cantidad_pagos"], 1)
        self.assertEqual(str(reporte.data["cobranzas"]["total"]), "12500")
        self.assertEqual(str(reporte.data["cobranzas"]["por_medio"][Pago.Medio.EFECTIVO]), "12500")
        self.assertEqual(str(reporte.data["cuenta_corriente"]["deuda"]), "20000.00")
        self.assertEqual(str(reporte.data["cuenta_corriente"]["saldo_a_favor"]), "12500.00")
        self.assertEqual(str(reporte.data["cuenta_corriente"]["saldo_neto"]), "7500.00")

    def test_payments_can_be_filtered_and_exported_as_csv(self):
        self.client.force_authenticate(user=self.admin)
        pedro = Alumno.objects.get(legajo="P-001")
        elena = Alumno.objects.get(legajo="E-001")
        Pago.objects.create(alumno=pedro, sucursal=self.posadas, importe="1000.00", medio=Pago.Medio.EFECTIVO)
        Pago.objects.create(alumno=elena, sucursal=self.eldorado, importe="2000.00", medio=Pago.Medio.TRANSFERENCIA)

        pagos = self.client.get(f"/api/pagos/?sucursal={self.posadas.id}&medio=efectivo")
        self.assertEqual(pagos.status_code, status.HTTP_200_OK)
        self.assertEqual(pagos.data["count"], 1)
        exportacion = self.client.get(f"/api/pagos/exportar-csv/?sucursal={self.posadas.id}")
        self.assertEqual(exportacion.status_code, status.HTTP_200_OK)
        self.assertEqual(exportacion["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("Perez, Pedro", exportacion.content.decode("utf-8-sig"))
        self.assertNotIn("Silva, Elena", exportacion.content.decode("utf-8-sig"))

    def test_debtors_calculate_debt_oldest_overdue_fee_and_last_payment(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        carrera = CarreraCurso.objects.get(nombre="Secretariado")
        alumno.carrera = carrera
        alumno.telefono = "3764001234"
        alumno.email = "pedro@example.com"
        alumno.save(update_fields=["carrera", "telefono", "email"])
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        hoy = timezone.localdate()
        vencida = Cuota.objects.create(
            alumno=alumno,
            concepto=concepto,
            sucursal=self.posadas,
            periodo="2026-05",
            fecha_emision=hoy - timedelta(days=60),
            fecha_vencimiento=hoy - timedelta(days=30),
            importe="30000.00",
            descuento="1000.00",
        )
        Cuota.objects.create(
            alumno=alumno,
            concepto=concepto,
            sucursal=self.posadas,
            periodo="2026-06",
            fecha_emision=hoy - timedelta(days=20),
            fecha_vencimiento=hoy + timedelta(days=5),
            importe="20000.00",
        )
        pago = Pago.objects.create(
            alumno=alumno,
            concepto=concepto,
            sucursal=self.posadas,
            importe="5000.00",
            medio=Pago.Medio.TRANSFERENCIA,
            fecha=hoy,
        )
        AplicacionPago.objects.create(pago=pago, cuota=vencida, importe="5000.00")

        response = self.client.get("/api/deudores/?sucursal=%s&deuda_min=40000" % self.posadas.id)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        debtor = response.data["results"][0]
        self.assertEqual(debtor["alumno"]["legajo"], "P-001")
        self.assertEqual(debtor["sucursal"], self.posadas.id)
        self.assertEqual(debtor["carrera"], carrera.id)
        self.assertEqual(debtor["deuda_total"], "44000.00")
        self.assertEqual(debtor["cuotas_pendientes"], 2)
        self.assertEqual(debtor["cuotas_vencidas"], 1)
        self.assertEqual(debtor["cuota_vencida_mas_antigua"], (hoy - timedelta(days=30)).isoformat())
        self.assertEqual(debtor["fecha_ultimo_pago"], hoy.isoformat())
        self.assertEqual(debtor["telefono"], "3764001234")

    def test_debtors_support_search_career_overdue_filter_and_branch_scope(self):
        self.client.force_authenticate(user=self.admin)
        pedro = Alumno.objects.get(legajo="P-001")
        concepto_posadas = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        carrera = CarreraCurso.objects.get(nombre="Secretariado")
        pedro.carrera = carrera
        pedro.save(update_fields=["carrera"])
        Cuota.objects.create(
            alumno=pedro,
            concepto=concepto_posadas,
            sucursal=self.posadas,
            periodo="2026-07",
            fecha_emision=timezone.localdate() - timedelta(days=20),
            fecha_vencimiento=timezone.localdate() - timedelta(days=1),
            importe="18000.00",
        )
        concepto_eldorado = ConceptoCobrable.objects.create(
            nombre="Cuota Eldorado",
            tipo=ConceptoCobrable.Tipo.CUOTA,
            importe="12000.00",
            sucursal=self.eldorado,
        )
        elena = Alumno.objects.get(legajo="E-001")
        Cuota.objects.create(
            alumno=elena,
            concepto=concepto_eldorado,
            sucursal=self.eldorado,
            periodo="2026-07",
            fecha_emision=timezone.localdate(),
            fecha_vencimiento=timezone.localdate(),
            importe="12000.00",
        )

        filtered = self.client.get(
            "/api/deudores/",
            {"search": "P-001", "carrera": carrera.id, "vencidas": "1", "orden": "antiguedad", "page_size": 5},
        )
        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.data["count"], 1)
        self.assertEqual(filtered.data["results"][0]["legajo"], "P-001")
        self.assertEqual(filtered.data["page_size"], 5)

        self.client.force_authenticate(user=self.cajero)
        scoped = self.client.get("/api/deudores/")
        self.assertEqual(scoped.status_code, status.HTTP_200_OK)
        self.assertTrue(all(item["sucursal"] == self.posadas.id for item in scoped.data["results"]))
        self.assertNotIn(elena.id, [item["id"] for item in scoped.data["results"]])

    def test_collections_by_user_excludes_voided_payments_and_breaks_down_methods(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        Pago.objects.create(
            alumno=alumno, sucursal=self.posadas, registrado_por=self.admin,
            importe="1200.00", medio=Pago.Medio.EFECTIVO,
        )
        Pago.objects.create(
            alumno=alumno, sucursal=self.posadas, registrado_por=self.admin,
            importe="800.00", medio=Pago.Medio.MERCADO_PAGO,
        )
        Pago.objects.create(
            alumno=alumno, sucursal=self.posadas, registrado_por=self.admin,
            importe="999.00", medio=Pago.Medio.EFECTIVO, estado=Pago.Estado.ANULADO,
        )

        response = self.client.get("/api/reportes/cobranzas-usuarios/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total"], Decimal("2000"))
        row = response.data["resultados"][0]
        self.assertEqual(row["usuario"], self.admin.username)
        self.assertEqual(row["efectivo"], Decimal("1200"))
        self.assertEqual(row["mercado_pago"], Decimal("800"))

    def test_excel_export_preserves_typed_payment_values(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        Pago.objects.create(
            alumno=alumno, sucursal=self.posadas, registrado_por=self.admin,
            importe="2500.50", medio=Pago.Medio.TRANSFERENCIA,
        )

        response = self.client.get("/api/reportes/exportar.xlsx?tipo=pagos")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(b"PK"))
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Recibo")
        self.assertEqual(sheet["I2"].value, 2500.5)

    def test_excel_export_supports_students_and_collection_user_filter(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        other = User.objects.create_user(username="otro-cajero", password="test")
        PerfilUsuario.objects.create(user=other, rol=PerfilUsuario.Rol.CAJA, sucursal=self.posadas)
        Pago.objects.create(
            alumno=alumno, sucursal=self.posadas, registrado_por=other,
            importe="700.00", medio=Pago.Medio.EFECTIVO,
        )

        filtered = self.client.get("/api/reportes/cobranzas-usuarios/", {"usuario": other.id})
        exported = self.client.get("/api/reportes/exportar.xlsx?tipo=alumnos&sucursal=%s" % self.posadas.id)

        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.data["resultados"][0]["usuario"], other.username)
        self.assertEqual(exported.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(exported.content), data_only=True)
        self.assertEqual(workbook.active["A1"].value, "Legajo")
        self.assertIn("P-001", [cell.value for cell in workbook.active["A"]])

    def test_configured_percentage_discount_is_calculated_and_traced(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        discount = TipoDescuento.objects.create(
            nombre="Beca",
            modalidad=TipoDescuento.Modalidad.PORCENTAJE,
            valor="20.00",
            sucursal=self.posadas,
        )

        response = self.client.post("/api/cuotas/", {
            "alumno": alumno.id,
            "concepto": concepto.id,
            "periodo": "09-2026",
            "fecha_emision": timezone.localdate(),
            "fecha_vencimiento": timezone.localdate(),
            "importe": "25000.00",
            "tipo_descuento": discount.id,
            "motivo_descuento": "Beca aprobada",
        }, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["descuento"], "5000.00")
        cuota = Cuota.objects.get(pk=response.data["id"])
        self.assertEqual(cuota.descuento_registrado_por, self.admin)
        self.assertEqual(cuota.total, Decimal("20000"))

    def test_overdue_surcharge_respects_tolerance_and_excludes_paid_fees(self):
        self.client.force_authenticate(user=self.admin)
        alumno = Alumno.objects.get(legajo="P-001")
        concepto = ConceptoCobrable.objects.get(nombre="Cuota mensual")
        rule = ReglaRecargo.objects.create(
            nombre="Mora 10%",
            sucursal=self.posadas,
            concepto=concepto,
            modalidad=ReglaRecargo.Modalidad.PORCENTAJE,
            valor="10.00",
            dias_tolerancia=5,
        )
        overdue = Cuota.objects.create(
            alumno=alumno, concepto=concepto, sucursal=self.posadas, periodo="MORA-1",
            fecha_emision=timezone.localdate() - timedelta(days=20),
            fecha_vencimiento=timezone.localdate() - timedelta(days=6), importe="10000.00",
        )
        boundary = Cuota.objects.create(
            alumno=alumno, concepto=concepto, sucursal=self.posadas, periodo="MORA-2",
            fecha_emision=timezone.localdate() - timedelta(days=20),
            fecha_vencimiento=timezone.localdate() - timedelta(days=5), importe="10000.00",
        )

        response = self.client.post("/api/reglas-recargo/recalcular/", {"sucursal": self.posadas.id}, format="json")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        overdue.refresh_from_db()
        boundary.refresh_from_db()
        self.assertEqual(overdue.recargo, Decimal("1000"))
        self.assertEqual(overdue.regla_recargo, rule)
        self.assertEqual(boundary.recargo, Decimal("0"))
class CobroEndpointTests(APITestCase):
    """Cubre POST /api/pagos/cobrar/ con aplicaciones manuales, automaticas, pago a cuenta, excedente y caja cerrada."""

    def setUp(self):
        self.posadas = Sucursal.objects.create(codigo="POS", nombre="Posadas")
        self.eldorado = Sucursal.objects.create(codigo="ELD", nombre="Eldorado")
        self.cajero = User.objects.create_user("cajero", password="cajero123")
        PerfilUsuario.objects.create(
            user=self.cajero,
            rol=PerfilUsuario.Rol.CAJA,
            sucursal=self.posadas,
        )
        self.alumno = Alumno.objects.create(
            legajo="POS-100",
            nombre="Lucia",
            apellido="Ramirez",
            dni="35111222",
            sucursal=self.posadas,
        )
        self.concepto = ConceptoCobrable.objects.create(
            nombre="Cuota mensual",
            tipo=ConceptoCobrable.Tipo.CUOTA,
            importe=10000,
            sucursal=self.posadas,
        )
        hoy = timezone.localdate()
        self.cuota_1 = Cuota.objects.create(
            alumno=self.alumno,
            concepto=self.concepto,
            sucursal=self.posadas,
            periodo="2026-07",
            fecha_emision=hoy,
            fecha_vencimiento=hoy,
            importe="10000.00",
        )
        self.cuota_2 = Cuota.objects.create(
            alumno=self.alumno,
            concepto=self.concepto,
            sucursal=self.posadas,
            periodo="2026-08",
            fecha_emision=hoy,
            fecha_vencimiento=hoy,
            importe="10000.00",
        )

    def _abrir_caja(self):
        self.client.force_authenticate(user=self.cajero)
        return self.client.get(f"/api/cajas/hoy/?sucursal={self.posadas.id}").data

    def test_cobro_con_aplicaciones_manuales_genera_pago_y_aplicaciones(self):
        caja = self._abrir_caja()
        self.assertEqual(caja["estado"], CajaDiaria.Estado.ABIERTA)

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "15000.00",
                "medio": Pago.Medio.EFECTIVO,
                "observacion": "Cobro manual",
                "aplicaciones": [
                    {"cuota_id": self.cuota_1.id, "importe": "10000.00"},
                    {"cuota_id": self.cuota_2.id, "importe": "5000.00"},
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.assertEqual(Pago.objects.count(), 1)
        self.assertEqual(AplicacionPago.objects.count(), 2)
        self.assertEqual(MovimientoCaja.objects.count(), 1)

        self.cuota_1.refresh_from_db()
        self.cuota_2.refresh_from_db()
        self.assertEqual(self.cuota_1.estado, Cuota.Estado.PAGADA)
        self.assertEqual(self.cuota_2.estado, Cuota.Estado.PARCIAL)

        self.assertEqual(str(response.data["importe_aplicado"]), "15000.00")
        self.assertEqual(str(response.data["saldo_a_favor"]), "0.00")
        self.assertEqual(len(response.data["aplicaciones"]), 2)
        self.assertTrue(response.data["numero_recibo"].startswith("REC-"))

    def test_cobro_modo_automatico_aplica_a_cuotas_mas_antiguas(self):
        self._abrir_caja()

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "12000.00",
                "medio": Pago.Medio.TRANSFERENCIA,
                "modo_automatico": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.cuota_1.refresh_from_db()
        self.cuota_2.refresh_from_db()
        self.assertEqual(self.cuota_1.estado, Cuota.Estado.PAGADA)
        self.assertEqual(self.cuota_2.estado, Cuota.Estado.PARCIAL)
        self.assertEqual(str(response.data["importe_aplicado"]), "12000.00")
        self.assertEqual(str(response.data["saldo_a_favor"]), "0.00")
        self.assertEqual(AplicacionPago.objects.filter(cuota=self.cuota_1).first().importe, Decimal("10000.00"))
        self.assertEqual(AplicacionPago.objects.filter(cuota=self.cuota_2).first().importe, Decimal("2000.00"))

    def test_cobro_pago_a_cuenta_sin_aplicaciones_genera_saldo_a_favor(self):
        self._abrir_caja()

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "5000.00",
                "medio": Pago.Medio.EFECTIVO,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.assertEqual(AplicacionPago.objects.count(), 0)
        self.assertEqual(str(response.data["saldo_a_favor"]), "5000.00")
        self.assertEqual(str(response.data["importe_aplicado"]), "0.00")

    def test_cobro_con_excedente_aplica_y_deja_saldo_a_favor(self):
        self._abrir_caja()

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "25000.00",
                "medio": Pago.Medio.EFECTIVO,
                "modo_automatico": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        self.assertEqual(str(response.data["importe_aplicado"]), "20000.00")
        self.assertEqual(str(response.data["saldo_a_favor"]), "5000.00")
        self.cuota_1.refresh_from_db()
        self.cuota_2.refresh_from_db()
        self.assertEqual(self.cuota_1.estado, Cuota.Estado.PAGADA)
        self.assertEqual(self.cuota_2.estado, Cuota.Estado.PAGADA)

    def test_cobro_rechaza_si_caja_esta_cerrada(self):
        caja_data = self._abrir_caja()
        self.client.post(
            f"/api/cajas/{caja_data['id']}/cerrar/",
            {"total_contado": "0.00"},
            format="json",
        )

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "10000.00",
                "medio": Pago.Medio.EFECTIVO,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("cerrada", response.data["detail"].lower())
        self.assertEqual(Pago.objects.count(), 0)

    def test_cobro_rechaza_aplicacion_que_supera_saldo_de_cuota(self):
        self._abrir_caja()

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "20000.00",
                "medio": Pago.Medio.EFECTIVO,
                "aplicaciones": [
                    {"cuota_id": self.cuota_1.id, "importe": "15000.00"},
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("aplicaciones", response.data)
        self.assertEqual(Pago.objects.count(), 0)
        self.assertEqual(AplicacionPago.objects.count(), 0)

    def test_cobro_rechaza_si_suma_de_aplicaciones_supera_importe(self):
        self._abrir_caja()

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "10000.00",
                "medio": Pago.Medio.EFECTIVO,
                "aplicaciones": [
                    {"cuota_id": self.cuota_1.id, "importe": "8000.00"},
                    {"cuota_id": self.cuota_2.id, "importe": "5000.00"},
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Pago.objects.count(), 0)

    def test_cobro_rechaza_si_cuota_es_de_otro_alumno(self):
        otro = Alumno.objects.create(
            legajo="POS-200",
            nombre="Mario",
            apellido="Suarez",
            dni="40111222",
            sucursal=self.posadas,
        )
        self._abrir_caja()

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": otro.id,
                "importe": "5000.00",
                "medio": Pago.Medio.EFECTIVO,
                "aplicaciones": [
                    {"cuota_id": self.cuota_1.id, "importe": "5000.00"},
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Pago.objects.count(), 0)

    def test_cobro_atomico_si_falla_no_genera_pago(self):
        self._abrir_caja()

        # Cuota con saldo suficiente, pero el monto es negativo: el serializer debe rechazarlo.
        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "-100.00",
                "medio": Pago.Medio.EFECTIVO,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Pago.objects.count(), 0)
        self.assertEqual(MovimientoCaja.objects.count(), 0)

    def test_cobro_actualiza_estado_cuenta_del_alumno(self):
        self._abrir_caja()

        response = self.client.post(
            "/api/pagos/cobrar/",
            {
                "alumno": self.alumno.id,
                "importe": "10000.00",
                "medio": Pago.Medio.EFECTIVO,
                "modo_automatico": True,
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)

        estado = self.client.get(f"/api/alumnos/{self.alumno.id}/estado-cuenta/")
        self.assertEqual(estado.status_code, status.HTTP_200_OK)
        self.assertEqual(str(estado.data["resumen"]["saldo_pendiente"]), "10000.00")
        self.assertEqual(str(estado.data["resumen"]["saldo_a_favor"]), "0.00")
        self.assertEqual(str(estado.data["resumen"]["saldo_neto"]), "10000.00")


# Create your tests here.

class AuditoriaApiTests(APITestCase):
    def setUp(self):
        self.posadas = Sucursal.objects.create(codigo="POS-AUD", nombre="Posadas auditoría")
        self.eldorado = Sucursal.objects.create(codigo="ELD-AUD", nombre="Eldorado auditoría")
        self.admin = User.objects.create_user(username="audit-admin", password="test")
        PerfilUsuario.objects.create(
            user=self.admin,
            rol=PerfilUsuario.Rol.ADMINISTRACION,
            sucursal=self.posadas,
        )
        self.consulta = User.objects.create_user(username="audit-read", password="test")
        PerfilUsuario.objects.create(
            user=self.consulta,
            rol=PerfilUsuario.Rol.CONSULTA,
            sucursal=self.posadas,
        )
        self.client.force_authenticate(self.admin)

    def test_create_student_registers_audit_event(self):
        response = self.client.post(
            "/api/alumnos/",
            {
                "legajo": "AUD-001",
                "nombre": "Ada",
                "apellido": "Lovelace",
                "dni": "40111222",
                "estado": Alumno.Estado.ACTIVO,
                "sucursal": self.posadas.id,
            },
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        event = EventoAuditoria.objects.get(entidad="core.Alumno", entidad_id=str(response.data["id"]))
        self.assertEqual(event.usuario, self.admin)
        self.assertEqual(event.sucursal, self.posadas)
        self.assertEqual(event.accion, "alta")
        self.assertEqual(event.valores_nuevos["legajo"], "AUD-001")

    def test_audit_log_is_read_only_scoped_and_not_available_to_consulta(self):
        EventoAuditoria.objects.create(
            usuario=self.admin,
            sucursal=self.posadas,
            modulo="alumnos",
            accion="alta",
            entidad="core.Alumno",
            entidad_id="1",
        )
        EventoAuditoria.objects.create(
            usuario=self.admin,
            sucursal=self.eldorado,
            modulo="alumnos",
            accion="alta",
            entidad="core.Alumno",
            entidad_id="2",
        )

        listed = self.client.get("/api/auditoria/?modulo=alumnos")
        self.assertEqual(listed.status_code, status.HTTP_200_OK)
        self.assertEqual(listed.data["count"], 1)
        self.assertEqual(listed.data["results"][0]["sucursal"], self.posadas.id)
        self.assertEqual(self.client.post("/api/auditoria/", {}).status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.consulta)
        self.assertEqual(self.client.get("/api/auditoria/").status_code, status.HTTP_403_FORBIDDEN)
